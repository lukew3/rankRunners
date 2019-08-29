# This program does the following:
# 1. Copy old sheet to new expendable sheet
# 2. Go through the entire list,
# 3. find the top runner
# 4. add him to the top of a new sheet
# 5. remove from old sheet
# 6. repeat
# note: eventually you should make functions for repeating sequences
#       also make a program that doesnt have the individual minutes and seconds cells
#       fix formatting of pr cell so that times are displayed like 19:09 instead of 19:9
#       make the program able to read off of the original file without any editing

import openpyxl
import os
if os.path.exists('D:\pythonProjects\pythonExcel\\rankedResults.xlsx'):
  os.remove('D:\pythonProjects\pythonExcel\\rankedResults.xlsx')
else:
  print("The file does not exist")

path = "D:\pythonProjects\pythonExcel\meetResultsSheet.xlsx"
#creates a duplicate sheet
wb = openpyxl.load_workbook(path)
sheet = wb.active
sheet.title = 'old'
wb.create_sheet(index=1, title='new')
wb.save('D:\pythonProjects\pythonExcel\\rankedResults.xlsx')

# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
oldSheet = wb['old']
newSheet = wb['new']

#variables are declared
rank = 1
rowNum = 3
tempBestRow = 0
tempBestMinutes = 100
tempBestSeconds = 100
rowMax = 46
count = 3

#functions
#nothing here yet

#Creates label row
rankTitleCell = newSheet.cell(row = 1, column = 1) 
firstNameTitleCell = newSheet.cell(row = 1, column = 2) 
lastNameTitleCell = newSheet.cell(row = 1, column = 3)
minutesTitleCell = newSheet.cell(row = 1, column = 4)
secondsTitleCell = newSheet.cell(row = 1, column = 5)
prTitleCell = newSheet.cell(row = 1, column = 6)
rankTitleCell.value = "Rank"
firstNameTitleCell = "First Name"
lastNameTitleCell = "Last Name"
minutesTitleCell = "Minutes"
secondsTitleCell = "Seconds"
prTitleCell = "PR Time"

while rank <= 44:
    while count <= rowMax:
        raceTimeMinutes = oldSheet.cell(row = rowNum, column = 3)
        raceTimeSeconds = oldSheet.cell(row = rowNum, column = 4)
        minutes = raceTimeMinutes.value
        seconds = raceTimeSeconds.value
        #print("Minutes are: " + str(minutes))
        #print("Seconds are: " + str(seconds))
        if minutes == tempBestMinutes:
            if seconds < tempBestSeconds:
                tempBestMinutes = minutes
                tempBestSeconds = seconds
                tempBestRow = rowNum
        elif minutes < tempBestMinutes:
            tempBestMinutes = minutes
            tempBestSeconds = seconds
            tempBestRow = rowNum
        rowNum = rowNum + 1
        count = count + 1
    
    #creates data that will be put into new cells from the old cells
    firstName = oldSheet.cell(row = tempBestRow, column = 1) 
    lastName = oldSheet.cell(row = tempBestRow, column = 2)
    first = firstName.value
    last = lastName.value
    raceTimeMinutes = oldSheet.cell(row = tempBestRow, column = 3)
    raceTimeSeconds = oldSheet.cell(row = tempBestRow, column = 4)
    raceTimeMinutesString = str(raceTimeMinutes.value)
    raceTimeSecondsString = str(raceTimeSeconds.value)
    raceTimeTotalString = raceTimeMinutesString + ":" + raceTimeSecondsString

    print(str(rank) + " " + first + " " + last + ": " + raceTimeMinutesString + ":" + raceTimeSecondsString)
    
    #defines cells going to be assigned
    rankCell = newSheet.cell(row = rank + 1, column = 1) 
    firstNameCell = newSheet.cell(row = rank + 1, column = 2) 
    lastNameCell = newSheet.cell(row = rank + 1, column = 3)
    minutesCell = newSheet.cell(row = rank + 1, column = 4)
    secondsCell = newSheet.cell(row = rank + 1, column = 5)
    totalPrCell = newSheet.cell(row = rank + 1, column = 6)
    
    #assigns values to cells
    rankCell.value = rank
    firstNameCell.value = first
    lastNameCell.value = last
    minutesCell.value = raceTimeMinutesString
    secondsCell.value = raceTimeSecondsString
    totalPrCell.value = raceTimeTotalString
    oldSheet.delete_rows(tempBestRow)
    rank = rank + 1

    #resets the variables that make the second function work 
    rowNum = 3
    tempBestRow = 0
    tempBestMinutes = 100
    tempBestSeconds = 100
    count = 3
    rowMax = rowMax-1

wb.remove(oldSheet)
wb.save('D:\pythonProjects\pythonExcel\\rankedResults.xlsx')